import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import io
import os
import json
from datetime import datetime
from PIL import Image as PILImage
import base64
import tempfile

class RAAWAGenerator:
    def __init__(self, db):
        self.db = db
        self.upload_folder = 'uploads'
        if not os.path.exists(self.upload_folder):
            os.makedirs(self.upload_folder)
    
    def generate_raawa_excel(self, raawa_id, raawa_data):
        """Generate Excel file for RAAWA"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "RAAWA Application"
            
            # Set column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 25
            
            # Styles
            title_font = Font(size=16, bold=True)
            header_font = Font(size=12, bold=True)
            label_font = Font(bold=True)
            center_alignment = Alignment(horizontal='center', vertical='center')
            border_style = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Title
            ws.merge_cells('A1:D1')
            ws['A1'] = "RAAWA APPLICATION"
            ws['A1'].font = title_font
            ws['A1'].alignment = center_alignment
            
            # Header border
            for row in range(1, 2):
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = border_style
            
            # RAAWA No
            row = 3
            ws[f'A{row}'] = "RAAWA No:"
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = raawa_data['raawa_no']
            ws[f'B{row}'].font = Font(bold=True, color="0000FF")
            ws.merge_cells(f'B{row}:D{row}')
            
            # Requisitioner Info
            row += 2
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'] = "REQUISITIONER INFORMATION"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].alignment = center_alignment
            ws[f'A{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            row += 1
            ws[f'A{row}'] = "Requisitioner Name:"
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = raawa_data['requisitioner_name']
            ws.merge_cells(f'B{row}:D{row}')
            
            row += 1
            ws[f'A{row}'] = "ID No.:"
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = raawa_data['id_no']
            ws.merge_cells(f'B{row}:D{row}')
            
            row += 1
            ws[f'A{row}'] = "Department/Group:"
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = raawa_data.get('department_group', '')
            ws.merge_cells(f'B{row}:D{row}')
            
            row += 1
            ws[f'A{row}'] = "Contact No.:"
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = raawa_data.get('contact_no', '')
            ws.merge_cells(f'B{row}:D{row}')
            
            row += 1
            ws[f'A{row}'] = "Region:"
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = raawa_data['region']
            ws.merge_cells(f'B{row}:D{row}')
            
            # Approvers
            row += 2
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'] = "APPROVERS"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].alignment = center_alignment
            ws[f'A{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            # Get approver names
            facility_manager = self.db.get_approver_name(raawa_data.get('facility_manager_id'))
            security = self.db.get_approver_name(raawa_data.get('security_id'))
            
            row += 1
            ws[f'A{row}'] = "Facility Manager:"
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = facility_manager
            ws.merge_cells(f'B{row}:D{row}')
            
            row += 1
            ws[f'A{row}'] = "Security:"
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = security
            ws.merge_cells(f'B{row}:D{row}')
            
            # Personnel List
            row += 2
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'] = "PERSONNEL LIST"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].alignment = center_alignment
            ws[f'A{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            row += 1
            headers = ['#', 'Name', 'Company', 'SEC ID']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = center_alignment
                cell.border = border_style
            
            # Add personnel data
            personnel = raawa_data.get('personnel', [])
            for idx, person in enumerate(personnel, 1):
                row += 1
                ws.cell(row=row, column=1, value=idx).border = border_style
                ws.cell(row=row, column=2, value=person['name']).border = border_style
                ws.cell(row=row, column=3, value=person.get('company', '')).border = border_style
                ws.cell(row=row, column=4, value=person.get('sec_id', '')).border = border_style
            
            # Footer
            row += 2
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws[f'A{row}'].alignment = Alignment(horizontal='right')
            
            # Save file
            filename = f"RAAWA_{raawa_data['raawa_no']}.xlsx"
            filepath = os.path.join(self.upload_folder, filename)
            wb.save(filepath)
            
            # Update database with file path
            self.db.update_raawa(raawa_id, {'file_path': filepath})
            
            return filepath
            
        except Exception as e:
            print(f"Error generating Excel: {e}")
            raise
    
    def generate_raawa_pdf(self, raawa_id):
        """Generate PDF for approved RAAWA"""
        try:
            raawa = self.db.get_raawa_by_id(raawa_id)
            if not raawa:
                raise Exception("RAAWA not found")
            
            personnel = self.db.get_raawa_personnel(raawa_id)
            
            # Create PDF
            filename = f"RAAWA_{raawa['raawa_no']}_approved.pdf"
            filepath = os.path.join(self.upload_folder, filename)
            
            doc = SimpleDocTemplate(filepath, pagesize=letter, 
                                   rightMargin=72, leftMargin=72,
                                   topMargin=72, bottomMargin=72)
            
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                alignment=1,  # Center
                spaceAfter=20
            )
            story.append(Paragraph("RAAWA APPLICATION - APPROVED", title_style))
            
            # RAAWA No
            raawa_style = ParagraphStyle(
                'RaawaStyle',
                parent=styles['Normal'],
                fontSize=14,
                textColor=colors.blue
            )
            story.append(Paragraph(f"RAAWA No: <b>{raawa['raawa_no']}</b>", raawa_style))
            story.append(Spacer(1, 0.2*inch))
            
            # Requisitioner Info
            story.append(Paragraph("REQUISITIONER INFORMATION", styles['Heading2']))
            info_data = [
                ['Requisitioner Name:', raawa['requisitioner_name']],
                ['ID No.:', raawa['id_no']],
                ['Department/Group:', raawa.get('department_group', '')],
                ['Contact No.:', raawa.get('contact_no', '')],
                ['Region:', raawa['region']]
            ]
            
            info_table = Table(info_data, colWidths=[2*inch, 3*inch])
            info_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ]))
            story.append(info_table)
            story.append(Spacer(1, 0.3*inch))
            
            # Approvers
            story.append(Paragraph("APPROVERS", styles['Heading2']))
            facility_manager = self.db.get_approver_name(raawa.get('facility_manager_id'))
            security = self.db.get_approver_name(raawa.get('security_id'))
            
            approver_data = [
                ['Facility Manager:', facility_manager, 
                 'Signed' if raawa.get('facility_manager_signature') else 'Pending'],
                ['Security:', security,
                 'Signed' if raawa.get('security_signature') else 'Pending']
            ]
            
            approver_table = Table(approver_data, colWidths=[2*inch, 2*inch, 1.5*inch])
            approver_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ]))
            story.append(approver_table)
            story.append(Spacer(1, 0.3*inch))
            
            # Personnel List
            story.append(Paragraph("PERSONNEL LIST", styles['Heading2']))
            
            if personnel:
                person_data = [['#', 'Name', 'Company', 'SEC ID']]
                for idx, person in enumerate(personnel, 1):
                    person_data.append([
                        str(idx),
                        person['name'],
                        person.get('company', ''),
                        person.get('sec_id', '')
                    ])
                
                person_table = Table(person_data, colWidths=[0.5*inch, 2.5*inch, 2*inch, 1.5*inch])
                person_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ]))
                story.append(person_table)
            else:
                story.append(Paragraph("No personnel listed", styles['Normal']))
            
            story.append(Spacer(1, 0.3*inch))
            
            # Signatures
            story.append(Paragraph("ELECTRONIC SIGNATURES", styles['Heading2']))
            
            # Add signature images if available
            signature_data = []
            
            if raawa.get('facility_manager_signature'):
                try:
                    sig_data = base64.b64decode(raawa['facility_manager_signature'])
                    sig_img = PILImage.open(io.BytesIO(sig_data))
                    # Resize signature
                    sig_img.thumbnail((150, 60))
                    sig_buffer = io.BytesIO()
                    sig_img.save(sig_buffer, format='PNG')
                    sig_buffer.seek(0)
                    signature_data.append(['Facility Manager:', RLImage(sig_buffer, width=1.5*inch, height=0.5*inch)])
                except:
                    signature_data.append(['Facility Manager:', 'Signature not available'])
            else:
                signature_data.append(['Facility Manager:', 'Not Signed'])
            
            if raawa.get('security_signature'):
                try:
                    sig_data = base64.b64decode(raawa['security_signature'])
                    sig_img = PILImage.open(io.BytesIO(sig_data))
                    sig_img.thumbnail((150, 60))
                    sig_buffer = io.BytesIO()
                    sig_img.save(sig_buffer, format='PNG')
                    sig_buffer.seek(0)
                    signature_data.append(['Security:', RLImage(sig_buffer, width=1.5*inch, height=0.5*inch)])
                except:
                    signature_data.append(['Security:', 'Signature not available'])
            else:
                signature_data.append(['Security:', 'Not Signed'])
            
            signature_table = Table(signature_data, colWidths=[2*inch, 3.5*inch])
            signature_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ]))
            story.append(signature_table)
            
            # Footer
            story.append(Spacer(1, 0.5*inch))
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.grey,
                alignment=1
            )
            story.append(Paragraph(
                f"ESig Reference No.: {raawa.get('esig_ref_no', 'N/A')} | "
                f"Approved on: {raawa.get('approved_at', 'N/A')}",
                footer_style
            ))
            story.append(Paragraph(
                f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                footer_style
            ))
            
            # Build PDF
            doc.build(story)
            
            # Update database with PDF path
            self.db.update_raawa(raawa_id, {'final_file_path': filepath})
            
            return filepath
            
        except Exception as e:
            print(f"Error generating PDF: {e}")
            raise
    
    def get_template_excel(self):
        """Get the template Excel file for RAAWA generation"""
        # This method is for reference - the actual template is generated dynamically
        pass
